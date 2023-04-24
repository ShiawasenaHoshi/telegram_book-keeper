import os

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from app.report.styles import BOLD, r_thick, thin, adjust_columns_size, top_thick, \
    r_bottom_thick, bottom_thick
from config import Config


def generate_report(data, sheet_name="Выписка"):
    wb = Workbook()
    del wb["Sheet"]
    titles = ["Дата", "Время", "Пользователь", "Сумма", "В валюте", "Описание", "Категория"]
    ws = wb.create_sheet(sheet_name)
    adjust_columns_size(ws)
    append_header(ws, titles)
    append_multiline(ws, data)
    FullRange = "A1:" + get_column_letter(ws.max_column) \
                + str(ws.max_row)
    ws.auto_filter.ref = FullRange
    return save_workbook(wb)


def save_workbook(workbook):
    path = os.path.join(Config.TEMP_FOLDER, f'report.xlsx')
    workbook.save(path)
    return path


def append_header(ws, titles):
    cells = []
    for i in range(len(titles)):
        cell = header_cell(ws, titles[i])
        cell.border = top_thick
        cell.alignment = Alignment(horizontal='center', vertical='bottom')
        if i == len(titles) - 1:
            cell.border = r_bottom_thick
        else:
            cell.border = bottom_thick
        cells.append(cell)

    ws.append(cells)


def header_cell(ws, data):
    cell = WriteOnlyCell(ws, value=data)
    cell.font = BOLD
    return cell


def cells_from_data(ws, row, last_row=False):
    cells = []
    last_index = len(row) - 1
    for i in range(0, len(row)):
        cell = WriteOnlyCell(ws, value=row[i])
        cell.alignment = Alignment(horizontal='center', vertical='bottom')
        if last_row:
            if i == last_index:
                cell.border = r_bottom_thick
            else:
                cell.border = bottom_thick
        else:
            if i == last_index:
                cell.border = r_thick
            else:
                cell.border = thin
        cells.append(cell)
    return cells


def append_multiline(ws, data):
    while len(data) > 1:
        ws.append(cells_from_data(ws, data.pop(0)))
    if len(data) >= 1:
        ws.append(cells_from_data(ws, data.pop(0), last_row=True))
